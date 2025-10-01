import os
import shutil
from datetime import datetime
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import current_app
from .models import db, Tenant, Property, Transaction

class DatabaseService:
    """Service class for database operations."""
    
    @staticmethod
    def backup_database():
        """Create a backup of the database file."""
        try:
            # Get the database path from the configured DATABASE_URI
            database_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
            
            # Handle SQLite database paths
            if database_uri.startswith('sqlite:///'):
                # Remove 'sqlite:///' prefix
                db_filename = database_uri.replace('sqlite:///', '')
                
                # Flask creates an 'instance' directory for the database
                # Construct the full path: app_root/instance/db_filename
                db_path = os.path.join(current_app.root_path, 'instance', db_filename)
            else:
                raise ValueError('Backup is only supported for SQLite databases')
            
            # Check if the database file exists
            if not os.path.exists(db_path):
                raise FileNotFoundError(f'Database file not found at: {db_path}')
            
            # Get backup storage path from environment variable
            backup_storage_path = current_app.config.get('BACKUP_STORAGE_PATH', '.')
            
            # Ensure the backup directory exists
            os.makedirs(backup_storage_path, exist_ok=True)
            
            # Create a dynamic filename for the backup
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"app_backup_{timestamp}.db"
            
            # Create the full backup file path
            backup_file_path = os.path.join(backup_storage_path, backup_filename)
            
            # Save a copy on the server's file system
            shutil.copy2(db_path, backup_file_path)
            
            return backup_file_path, backup_filename
            
        except Exception as e:
            raise Exception(f"Backup failed: {str(e)}")

class ReportService:
    """Service class for generating reports."""
    
    @staticmethod
    def generate_csv_report(data, headers):
        """Generate a CSV report from data and headers."""
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)
        output.seek(0)
        return BytesIO(output.getvalue().encode('utf-8'))
    
    @staticmethod
    def generate_excel_report(data, headers, sheet_name="Report"):
        """Generate an Excel report from data and headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

class TenantService:
    """Service class for tenant operations."""
    
    @staticmethod
    def get_all_tenants():
        """Get all tenants with pagination."""
        return Tenant.query.all()
    
    @staticmethod
    def get_tenant_by_id(tenant_id):
        """Get a tenant by ID."""
        return Tenant.query.get_or_404(tenant_id)
    
    @staticmethod
    def create_tenant(data):
        """Create a new tenant."""
        tenant = Tenant(**data)
        db.session.add(tenant)
        db.session.commit()
        return tenant
    
    @staticmethod
    def update_tenant(tenant_id, data):
        """Update an existing tenant."""
        tenant = TenantService.get_tenant_by_id(tenant_id)
        for key, value in data.items():
            if hasattr(tenant, key):
                setattr(tenant, key, value)
        db.session.commit()
        return tenant
    
    @staticmethod
    def delete_tenant(tenant_id):
        """Delete a tenant."""
        tenant = TenantService.get_tenant_by_id(tenant_id)
        db.session.delete(tenant)
        db.session.commit()
        return tenant

class PropertyService:
    """Service class for property operations."""
    
    @staticmethod
    def get_all_properties():
        """Get all properties."""
        return Property.query.all()
    
    @staticmethod
    def get_property_by_id(property_id):
        """Get a property by ID."""
        return Property.query.get_or_404(property_id)
    
    @staticmethod
    def create_property(data):
        """Create a new property."""
        property_obj = Property(**data)
        db.session.add(property_obj)
        db.session.commit()
        return property_obj
    
    @staticmethod
    def update_property(property_id, data):
        """Update an existing property."""
        property_obj = PropertyService.get_property_by_id(property_id)
        for key, value in data.items():
            if hasattr(property_obj, key):
                setattr(property_obj, key, value)
        db.session.commit()
        return property_obj
    
    @staticmethod
    def delete_property(property_id):
        """Delete a property."""
        property_obj = PropertyService.get_property_by_id(property_id)
        db.session.delete(property_obj)
        db.session.commit()
        return property_obj

class TransactionService:
    """Service class for transaction operations."""
    
    @staticmethod
    def get_all_transactions():
        """Get all transactions."""
        return Transaction.query.all()
    
    @staticmethod
    def get_transaction_by_id(transaction_id):
        """Get a transaction by ID."""
        return Transaction.query.get_or_404(transaction_id)
    
    @staticmethod
    def create_transaction(data):
        """Create a new transaction."""
        transaction = Transaction(**data)
        db.session.add(transaction)
        db.session.commit()
        return transaction
    
    @staticmethod
    def update_transaction(transaction_id, data):
        """Update an existing transaction."""
        transaction = TransactionService.get_transaction_by_id(transaction_id)
        for key, value in data.items():
            if hasattr(transaction, key):
                setattr(transaction, key, value)
        db.session.commit()
        return transaction
    
    @staticmethod
    def delete_transaction(transaction_id):
        """Delete a transaction."""
        transaction = TransactionService.get_transaction_by_id(transaction_id)
        db.session.delete(transaction)
        db.session.commit()
        return transaction
