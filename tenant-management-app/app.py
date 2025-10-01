# app.py - A single-file, full-stack property management application.

import os
from datetime import datetime, date, timedelta
from flask import Flask, render_template_string, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO, StringIO
import csv
from dotenv import load_dotenv
import shutil

# Load environment variables from a .env file. This must be called before
# any os.getenv() calls that rely on the .env file.
load_dotenv()

# --- Embedded Frontend HTML/CSS/JS ---
# This is the entire user interface, a single HTML template that Flask will serve.
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Property Management</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        /* Custom styles for a clean, modern look */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;700&display=swap');
        body {
            font-family: 'Inter', sans-serif;
            background-color: #f3f4f6;
        }
        .tab-button {
            transition: all 0.2s;
            font-weight: 500;
        }
        .tab-button.active {
            border-bottom-width: 2px;
            border-color: #3b82f6;
            color: #3b82f6;
        }
        .main-container {
            max-width: 1200px;
        }
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.75);
            display: flex;
            justify-content: center;
            align-items: center;
            z-index: 1000;
            transition: opacity 0.3s ease-in-out;
            opacity: 0;
            pointer-events: none;
        }
        .modal-overlay.visible {
            opacity: 1;
            pointer-events: auto;
        }
        .modal-content {
            background-color: #fff;
            padding: 2rem;
            border-radius: 1rem;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.2);
            max-width: 90%;
            max-height: 90vh;
            overflow-y: auto;
            transform: scale(0.9);
            transition: transform 0.3s ease-in-out;
        }
        .modal-overlay.visible .modal-content {
            transform: scale(1);
        }
        .id-link {
            cursor: pointer;
            color: #3b82f6;
            text-decoration: underline;
        }
        .sortable-header {
            cursor: pointer;
            user-select: none;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .sortable-header:hover {
            color: #1e40af;
        }
        .sort-icon {
            width: 1rem;
            height: 1rem;
            margin-left: 0.25rem;
        }
        .pagination-button.active {
            background-color: #3b82f6;
            color: #fff;
        }
        .pagination-button:disabled {
            cursor: not-allowed;
            opacity: 0.5;
        }
    </style>
</head>
<body class="bg-gray-100 flex items-center justify-center min-h-screen p-4">
    <div class="bg-white shadow-xl rounded-2xl w-full main-container">
        <!-- Header -->
        <header class="p-6 border-b border-gray-200">
            <h1 class="text-3xl font-bold text-gray-800">Property Management</h1>
            <p class="text-gray-500 mt-1">Manage tenants, properties, and transactions with ease.</p>
        </header>

        <!-- Tabs Navigation -->
        <nav class="flex border-b border-gray-200 text-sm md:text-base px-6">
            <button id="tab-transactions" class="tab-button py-4 px-4 text-gray-600 active focus:outline-none">Transactions</button>
            <button id="tab-tenants" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Tenants</button>
            <button id="tab-properties" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Properties</button>
            <button id="tab-reports" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Reports</button>
        </nav>

        <main class="p-6">
            <!-- Transactions Section -->
            <section id="section-transactions" class="tab-content">
                <div class="flex flex-col md:flex-row justify-between items-start md:items-center mb-6 space-y-4 md:space-y-0">
                    <h2 class="text-xl font-bold text-gray-800">Transactions</h2>
                    <div class="flex flex-col sm:flex-row space-y-4 sm:space-y-0 sm:space-x-4">
                        <div>
                            <label for="transaction-property-filter" class="text-sm font-medium text-gray-700 mr-2">Filter by Property:</label>
                            <select id="transaction-property-filter" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="">All</option>
                            </select>
                        </div>
                        <div>
                            <label for="transaction-type-filter" class="text-sm font-medium text-gray-700 mr-2">Filter by Type:</label>
                            <select id="transaction-type-filter" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="all">All</option>
                                <option value="rent">Rent</option>
                                <option value="security">Security</option>
                                <option value="payment_received">Payment Received</option>
                                <option value="gas">Gas</option>
                                <option value="electricity">Electricity</option>
                                <option value="water">Water</option>
                                <option value="maintenance">Maintenance</option>
                                <option value="misc">Miscellaneous</option>
                            </select>
                        </div>
                        <button id="export-transactions-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-transaction-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Transaction</button>
                    </div>
                </div>
                <div id="transaction-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="transaction-form" class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4">
                        <input type="hidden" id="transaction-id">
                        <div>
                            <label for="transaction-tenant" class="block text-sm font-medium text-gray-700">Tenant</label>
                            <select id="transaction-tenant" name="tenant_id" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></select>
                        </div>
                        <div>
                            <label for="transaction-property" class="block text-sm font-medium text-gray-700">Property</label>
                            <select id="transaction-property" disabled class="mt-1 block w-full rounded-md border-gray-300 shadow-sm bg-gray-200 cursor-not-allowed"></select>
                            <input type="hidden" id="hidden-transaction-property-id" name="property_id">
                        </div>
                        <div>
                            <label for="transaction-type" class="block text-sm font-medium text-gray-700">Type</label>
                            <select id="transaction-type" name="type" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="rent">Rent</option>
                                <option value="security">Security</option>
                                <option value="payment_received">Payment Received</option>
                                <option value="gas">Gas</option>
                                <option value="electricity">Electricity</option>
                                <option value="water">Water</option>
                                <option value="maintenance">Maintenance</option>
                                <option value="misc">Miscellaneous</option>
                            </select>
                        </div>
                        <div>
                            <label for="transaction-for-month" class="block text-sm font-medium text-gray-700">For Month</label>
                            <select id="transaction-for-month" name="for_month" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="">Select Month</option>
                                <option value="January">January</option>
                                <option value="February">February</option>
                                <option value="March">March</option>
                                <option value="April">April</option>
                                <option value="May">May</option>
                                <option value="June">June</option>
                                <option value="July">July</option>
                                <option value="August">August</option>
                                <option value="September">September</option>
                                <option value="October">October</option>
                                <option value="November">November</option>
                                <option value="December">December</option>
                            </select>
                        </div>
                        <div>
                            <label for="transaction-amount" class="block text-sm font-medium text-gray-700">Amount</label>
                            <input type="number" id="transaction-amount" name="amount" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="transaction-transaction-date" class="block text-sm font-medium text-gray-700">Transaction Date</label>
                            <input type="date" id="transaction-transaction-date" name="transaction_date" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2">
                            <label for="transaction-comments" class="block text-sm font-medium text-gray-700">Comments</label>
                            <textarea id="transaction-comments" name="comments" rows="3" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></textarea>
                        </div>
                        <div class="col-span-1 md:col-span-2 lg:col-span-3 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-transaction-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Transaction</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="id">
                                        <span>ID</span>
                                        <span class="sort-icon" id="sort-icon-id"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Property</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Tenant</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="type">
                                        <span>Type</span>
                                        <span class="sort-icon" id="sort-icon-type"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">For Month</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="amount">
                                        <span>Amount</span>
                                        <span class="sort-icon" id="sort-icon-amount"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="transaction_date">
                                        <span>Transaction Date</span>
                                        <span class="sort-icon" id="sort-icon-transaction_date"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Comments</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="transaction-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Transaction data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
                <div id="pagination-controls" class="flex justify-center items-center space-x-2 mt-6">
                    <button id="prev-page" class="pagination-button bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Previous</button>
                    <div id="page-numbers" class="flex space-x-1"></div>
                    <button id="next-page" class="pagination-button bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Next</button>
                </div>
            </section>
            
            <!-- Tenants Section -->
            <section id="section-tenants" class="tab-content hidden">
                <div class="flex justify-between items-center mb-6">
                    <h2 class="text-xl font-bold text-gray-800">Tenants</h2>
                    <div class="flex space-x-2">
                        <button id="export-tenants-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-tenant-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Tenant</button>
                    </div>
                </div>
                <div id="tenant-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="tenant-form" class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4">
                        <input type="hidden" id="tenant-id">
                        <div>
                            <label for="tenant-name" class="block text-sm font-medium text-gray-700">Name</label>
                            <input type="text" id="tenant-name" name="name" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-property" class="block text-sm font-medium text-gray-700">Property</label>
                            <select id="tenant-property" name="property_id" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></select>
                        </div>
                        <div>
                            <label for="tenant-passport" class="block text-sm font-medium text-gray-700">Passport</label>
                            <input type="text" id="tenant-passport" name="passport" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-passport-validity" class="block text-sm font-medium text-gray-700">Passport Validity</label>
                            <input type="date" id="tenant-passport-validity" name="passport_validity" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-aadhar" class="block text-sm font-medium text-gray-700">Aadhar No</label>
                            <input type="text" id="tenant-aadhar" name="aadhar_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-employment" class="block text-sm font-medium text-gray-700">Employment Details</label>
                            <input type="text" id="tenant-employment" name="employment_details" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-address" class="block text-sm font-medium text-gray-700">Permanent Address</label>
                            <input type="text" id="tenant-address" name="permanent_address" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contact" class="block text-sm font-medium text-gray-700">Contact No</label>
                            <input type="text" id="tenant-contact" name="contact_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-emergency-contact" class="block text-sm font-medium text-gray-700">Emergency Contact No</label>
                            <input type="text" id="tenant-emergency-contact" name="emergency_contact_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-rent" class="block text-sm font-medium text-gray-700">Rent</label>
                            <input type="number" id="tenant-rent" name="rent" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-security" class="block text-sm font-medium text-gray-700">Security Deposit</label>
                            <input type="number" id="tenant-security" name="security" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-move-in-date" class="block text-sm font-medium text-gray-700">Move In Date</label>
                            <input type="date" id="tenant-move-in-date" name="move_in_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contract-start-date" class="block text-sm font-medium text-gray-700">Contract Start Date</label>
                            <input type="date" id="tenant-contract-start-date" name="contract_start_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contract-expiry-date" class="block text-sm font-medium text-gray-700">Contract Expiry Date</label>
                            <input type="date" id="tenant-contract-expiry-date" name="contract_expiry_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2 lg:col-span-3 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-tenant-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Tenant</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">ID</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Name</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Property</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Contact No</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Rent</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Contract Expiry Date</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="tenant-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Tenant data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
            </section>

            <!-- Properties Section -->
            <section id="section-properties" class="tab-content hidden">
                <div class="flex justify-between items-center mb-6">
                    <h2 class="text-xl font-bold text-gray-800">Properties</h2>
                    <div class="flex space-x-2">
                        <button id="export-properties-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-property-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Property</button>
                    </div>
                </div>
                <div id="property-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="property-form" class="grid grid-cols-1 md:grid-cols-2 gap-4">
                        <input type="hidden" id="property-id">
                        <div>
                            <label for="property-address" class="block text-sm font-medium text-gray-700">Address</label>
                            <input type="text" id="property-address" name="address" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="property-rent" class="block text-sm font-medium text-gray-700">Rent</label>
                            <input type="number" id="property-rent" name="rent" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="property-maintenance" class="block text-sm font-medium text-gray-700">Maintenance</label>
                            <input type="number" id="property-maintenance" name="maintenance" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-property-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Property</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">ID</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Address</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Rent</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Maintenance</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="property-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Property data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
            </section>

            <!-- Reports Section -->
            <section id="section-reports" class="tab-content hidden">
                <h2 class="text-xl font-bold text-gray-800 mb-4">Reports</h2>
                <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
                    <div class="bg-gray-50 p-6 rounded-xl shadow-sm flex flex-col items-start">
                        <h3 class="text-lg font-semibold text-gray-800">All Tenants Report (xlsx)</h3>
                        <p class="text-gray-600 mt-2">Generate a detailed report of all tenant information in an Excel file.</p>
                        <a href="/api/reports/tenants" class="mt-4 bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 self-end">Download .xlsx</a>
                    </div>
                    <div class="bg-gray-50 p-6 rounded-xl shadow-sm flex flex-col items-start">
                        <h3 class="text-lg font-semibold text-gray-800">All Transactions Report (xlsx)</h3>
                        <p class="text-gray-600 mt-2">Generate a detailed report of all transactions in an Excel file.</p>
                        <a href="/api/reports/transactions" class="mt-4 bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 self-end">Download .xlsx</a>
                    </div>
                    <div class="bg-gray-50 p-6 rounded-xl shadow-sm flex flex-col items-start md:col-span-2">
                        <h3 class="text-lg font-semibold text-gray-800">Database Backup</h3>
                        <p class="text-gray-600 mt-2">Download a full backup of your application's database file. A copy is also saved on the server.</p>
                        <a href="/api/backup" class="mt-4 bg-yellow-500 hover:bg-yellow-600 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 self-end">Backup & Download</a>
                    </div>
                </div>
            </section>
        </main>
    </div>

    <!-- Tenant Details Modal -->
    <div id="tenant-details-modal" class="modal-overlay">
        <div class="modal-content">
            <div class="flex justify-between items-center pb-4 border-b mb-4">
                <h3 class="text-2xl font-bold text-gray-800">Tenant Details</h3>
                <button id="close-tenant-modal-btn" class="text-gray-400 hover:text-gray-700 text-3xl">&times;</button>
            </div>
            <div id="tenant-details-content" class="grid grid-cols-1 md:grid-cols-2 gap-4 text-gray-700">
                <!-- Details will be populated here by JavaScript -->
            </div>
            <div class="flex justify-end mt-6">
                <button id="close-tenant-modal-footer-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Close</button>
            </div>
        </div>
    </div>
    
    <!-- Property Transactions Modal -->
    <div id="property-transactions-modal" class="modal-overlay">
        <div class="modal-content w-full max-w-4xl">
            <div class="flex justify-between items-center pb-4 border-b mb-4">
                <h3 class="text-2xl font-bold text-gray-800">Transactions for Property #<span id="prop-id-in-modal"></span></h3>
                <button id="close-prop-modal-btn" class="text-gray-400 hover:text-gray-700 text-3xl">&times;</button>
            </div>
            <div class="overflow-x-auto">
                <table class="min-w-full divide-y divide-gray-200">
                    <thead class="bg-gray-50">
                        <tr>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Tenant</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Type</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">For Month</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Amount</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Transaction Date</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Comments</th>
                        </tr>
                    </thead>
                    <tbody id="transactions-modal-table-body" class="bg-white divide-y divide-gray-200">
                        <!-- Transaction data will be populated here by JavaScript -->
                    </tbody>
                </table>
            </div>
            <div id="transactions-modal-footer" class="mt-6 flex justify-between items-center">
                <div class="text-lg font-bold text-gray-800">
                    Total Balance: <span id="transactions-modal-total" class=""></span>
                </div>
                <button id="close-prop-modal-footer-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Close</button>
            </div>
        </div>
    </div>

    <script>
        document.addEventListener('DOMContentLoaded', () => {
            const tabs = document.querySelectorAll('.tab-button');
            const sections = document.querySelectorAll('.tab-content');
            
            // Map plural model names to singular for consistent ID access
            const singularModels = {
                'tenants': 'tenant',
                'properties': 'property',
                'transactions': 'transaction'
            };

            const forms = {
                'tenants': document.getElementById('tenant-form-container'),
                'properties': document.getElementById('property-form-container'),
                'transactions': document.getElementById('transaction-form-container'),
            };
            const formButtons = {
                'tenants': document.getElementById('add-tenant-btn'),
                'properties': document.getElementById('add-property-btn'),
                'transactions': document.getElementById('add-transaction-btn'),
            };
            const tables = {
                'tenants': document.getElementById('tenant-table-body'),
                'properties': document.getElementById('property-table-body'),
                'transactions': document.getElementById('transaction-table-body'),
            };
            
            // Global data caches and state variables
            let allProperties = [];
            let allTenants = [];
            let allTransactions = [];
            
            // State for the transactions page
            let currentPage = 1;
            const itemsPerPage = 50;
            let totalPages = 1;
            let filterType = 'all';
            let filterPropertyId = '';
            let sortKey = 'transaction_date';
            let sortDirection = 'desc';

            // Default active tab is now 'transactions'
            let currentActiveTab = 'transactions';

            // Function to handle tab switching and show/hide sections
            function showSection(sectionId) {
                sections.forEach(section => {
                    section.classList.add('hidden');
                });
                document.getElementById(`section-${sectionId}`).classList.remove('hidden');

                tabs.forEach(tab => {
                    tab.classList.remove('active', 'text-blue-600', 'border-blue-600');
                    tab.classList.add('text-gray-600');
                });
                document.getElementById(`tab-${sectionId}`).classList.add('active', 'text-blue-600', 'border-blue-600');
                currentActiveTab = sectionId;
            }

            // Event listeners for tab buttons
            tabs.forEach(tab => {
                tab.addEventListener('click', () => {
                    const sectionId = tab.id.split('-')[1];
                    showSection(sectionId);
                    if (sectionId !== 'reports') {
                        if (sectionId === 'transactions') {
                            currentPage = 1;
                            fetchTransactions();
                        } else {
                            fetchData(sectionId);
                        }
                    }
                });
            });
            
            // Initial load of data for all sections
            showSection('transactions');
            fetchData('tenants');
            fetchData('properties');
            fetchTransactions();

            // --- API Functions ---
            // Generic function to fetch data for a given model (tenants, properties)
            async function fetchData(model) {
                const response = await fetch(`/api/${model}`);
                const data = await response.json();
                
                if (model === 'tenants') {
                    allTenants = data;
                    renderTable(model, data);
                    populateSelect('transaction-tenant', allTenants, 'name', 'id', '-- Select Tenant --');
                } else if (model === 'properties') {
                    allProperties = data;
                    renderTable(model, data);
                    populateSelect('transaction-property-filter', allProperties, 'address', 'id', 'All');
                    populateSelect('transaction-property', allProperties, 'address', 'id', '-- Select Property --');
                    populateSelect('tenant-property', allProperties, 'address', 'id', '-- Select Property --');
                } else {
                    renderTable(model, data);
                }
            }
            
            // Function to fetch paginated and filtered transactions
            async function fetchTransactions() {
                // Ensure properties and tenants are loaded before transactions
                if (allProperties.length === 0) {
                    await fetchData('properties');
                }
                if (allTenants.length === 0) {
                    await fetchData('tenants');
                }
                
                const url = `/api/transactions?page=${currentPage}&per_page=${itemsPerPage}&type=${filterType}&property_id=${filterPropertyId}&sort_by=${sortKey}&sort_direction=${sortDirection}`;
                const response = await fetch(url);
                const data = await response.json();
                
                allTransactions = data.items;
                totalPages = data.total_pages;
                
                renderTransactionsTable();
                renderPaginationControls();
            }
            
            // Generic function to create a new record
            async function postData(model, data) {
                const response = await fetch(`/api/${model}`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                if (response.ok) {
                    forms[model].classList.add('hidden');
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error creating record: ${error.error}`);
                }
            }

            // Generic function to update an existing record
            async function putData(model, id, data) {
                const response = await fetch(`/api/${model}/${id}`, {
                    method: 'PUT',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                if (response.ok) {
                    forms[model].classList.add('hidden');
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error updating record: ${error.error}`);
                }
            }

            // Generic function to delete a record
            async function deleteData(model, id) {
                if (!confirm(`Are you sure you want to delete this ${singularModels[model]}?`)) return;
                
                const response = await fetch(`/api/${model}/${id}`, {
                    method: 'DELETE'
                });
                if (response.ok) {
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error deleting record: ${error.error}`);
                }
            }

            // Fetches and displays a single tenant's detailed information in a modal
            async function showTenantDetails(id) {
                try {
                    const response = await fetch(`/api/tenants/${id}`);
                    if (!response.ok) throw new Error('Tenant not found.');
                    const tenant = await response.json();
                    
                    const detailsContent = document.getElementById('tenant-details-content');
                    detailsContent.innerHTML = `
                        <div class="col-span-1 md:col-span-2">
                            <p class="font-semibold">Name:</p>
                            <p>${tenant.name}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Property:</p>
                            <p>${tenant.property_address || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Passport:</p>
                            <p>${tenant.passport || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Passport Validity:</p>
                            <p>${tenant.passport_validity || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Aadhar No:</p>
                            <p>${tenant.aadhar_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Employment Details:</p>
                            <p>${tenant.employment_details || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Permanent Address:</p>
                            <p>${tenant.permanent_address || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contact No:</p>
                            <p>${tenant.contact_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Emergency Contact No:</p>
                            <p>${tenant.emergency_contact_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Rent:</p>
                            <p>$${(tenant.rent || 0).toFixed(2)}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Security Deposit:</p>
                            <p>$${(tenant.security || 0).toFixed(2)}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Move In Date:</p>
                            <p>${tenant.move_in_date || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contract Start Date:</p>
                            <p>${tenant.contract_start_date || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contract Expiry Date:</p>
                            <p>${tenant.contract_expiry_date || 'N/A'}</p>
                        </div>
                    `;
                    document.getElementById('tenant-details-modal').classList.add('visible');
                } catch (error) {
                    alert('Error fetching tenant details.');
                }
            }

            // Fetches and displays all transactions for a specific property in a modal
            async function showPropertyTransactions(id) {
                try {
                    const response = await fetch(`/api/properties/${id}/transactions`);
                    if (!response.ok) throw new Error('Could not fetch transactions.');
                    const { transactions, total } = await response.json();
                    
                    document.getElementById('prop-id-in-modal').textContent = id;
                    const transactionsTableBody = document.getElementById('transactions-modal-table-body');
                    transactionsTableBody.innerHTML = '';
                    
                    if (transactions.length === 0) {
                        transactionsTableBody.innerHTML = '<tr><td colspan="6" class="text-center py-4 text-gray-500">No transactions found for this property.</td></tr>';
                    } else {
                        transactions.forEach(tx => {
                            const date = new Date(tx.transaction_date).toLocaleDateString();
                            const displayAmount = (tx.type === 'payment_received') ? `$${tx.amount.toFixed(2)}` : `-$${tx.amount.toFixed(2)}`;
                            transactionsTableBody.innerHTML += `
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.tenant_name || 'N/A'}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.type}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.for_month || 'N/A'}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${displayAmount}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${date}</td>
                                    <td class="px-6 py-4 text-sm text-gray-500">${tx.comments || 'N/A'}</td>
                                </tr>
                            `;
                        });
                    }
                    
                    const totalElement = document.getElementById('transactions-modal-total');
                    totalElement.textContent = `$${total.toFixed(2)}`;
                    
                    totalElement.classList.remove('text-green-600', 'text-red-600', 'text-blue-600');
                    if (total >= 0) {
                        totalElement.classList.add('text-green-600');
                    } else {
                        totalElement.classList.add('text-red-600');
                    }
                    
                    document.getElementById('property-transactions-modal').classList.add('visible');
                } catch (error) {
                    alert('Error fetching property transactions.');
                }
            }
            
            // --- Rendering Functions ---
            // Renders the main table for tenants and properties
            function renderTable(model, data) {
                tables[model].innerHTML = '';
                let rowsHtml = '';
                if (model === 'tenants') {
                    const isExpiringSoon = (expiryDateStr) => {
                        if (!expiryDateStr) return false;
                        const twoMonthsFromNow = new Date();
                        twoMonthsFromNow.setMonth(twoMonthsFromNow.getMonth() + 2);
                        const expiryDate = new Date(expiryDateStr);
                        return expiryDate > new Date() && expiryDate < twoMonthsFromNow;
                    };

                    data.forEach(item => {
                        const dateClass = isExpiringSoon(item.contract_expiry_date) ? 'text-red-600 font-bold' : '';
                        rowsHtml += `
                            <tr data-id="${item.id}">
                                <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">
                                    <button class="id-link tenant-id-link" data-id="${item.id}">${item.id}</button>
                                </td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.name}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.property_address || 'N/A'}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.contact_no || 'N/A'}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.rent.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm ${dateClass}">${item.contract_expiry_date || 'N/A'}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                    <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                    <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                                </td>
                            </tr>
                        `;
                    });
                } else if (model === 'properties') {
                    data.forEach(item => {
                        rowsHtml += `
                            <tr data-id="${item.id}">
                                <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">
                                    <button class="id-link property-id-link" data-id="${item.id}">${item.id}</button>
                                </td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.address}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.rent.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.maintenance.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                    <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                    <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                                </td>
                            </tr>
                        `;
                    });
                }
                tables[model].innerHTML = rowsHtml;
            }

            // Renders the transactions table with the current page and filters
            function renderTransactionsTable() {
                tables['transactions'].innerHTML = '';
                let rowsHtml = '';
                allTransactions.forEach(item => {
                    const date = new Date(item.transaction_date).toLocaleDateString();
                    const displayAmount = (item.type === 'payment_received') ? `$${item.amount.toFixed(2)}` : `-$${item.amount.toFixed(2)}`;
                    rowsHtml += `
                        <tr data-id="${item.id}">
                            <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">${item.id}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.property_address || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.tenant_name || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.type}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.for_month || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${displayAmount}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${date}</td>
                            <td class="px-6 py-4 text-sm text-gray-500">${item.comments || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                            </td>
                        </tr>
                    `;
                });
                tables['transactions'].innerHTML = rowsHtml;
                updateSortIcons();
            }

            // Updates the sorting icons in the table headers
            function updateSortIcons() {
                document.querySelectorAll('.sort-icon').forEach(icon => {
                    icon.innerHTML = '';
                });
                if (sortKey) {
                    const iconElement = document.getElementById(`sort-icon-${sortKey}`);
                    if (iconElement) {
                        iconElement.innerHTML = sortDirection === 'asc' ? '&#9650;' : '&#9660;';
                    }
                }
            }
            
            // Renders the pagination buttons
            function renderPaginationControls() {
                const paginationDiv = document.getElementById('pagination-controls');
                const pageNumbersDiv = document.getElementById('page-numbers');
                const prevButton = document.getElementById('prev-page');
                const nextButton = document.getElementById('next-page');
                
                pageNumbersDiv.innerHTML = '';
                
                prevButton.disabled = currentPage === 1;
                nextButton.disabled = currentPage === totalPages || totalPages === 0;

                const startPage = Math.max(1, currentPage - 2);
                const endPage = Math.min(totalPages, currentPage + 2);
                
                if (startPage > 1) {
                    pageNumbersDiv.innerHTML += `<button class="pagination-button py-2 px-3 rounded-full hover:bg-gray-300">1</button><span>...</span>`;
                }

                for (let i = startPage; i <= endPage; i++) {
                    const button = document.createElement('button');
                    button.textContent = i;
                    button.classList.add('pagination-button', 'py-2', 'px-3', 'rounded-full', 'hover:bg-gray-300');
                    if (i === currentPage) {
                        button.classList.add('active');
                    }
                    button.addEventListener('click', () => {
                        currentPage = i;
                        fetchTransactions();
                    });
                    pageNumbersDiv.appendChild(button);
                }
                
                if (endPage < totalPages) {
                    pageNumbersDiv.innerHTML += `<span>...</span><button class="pagination-button py-2 px-3 rounded-full hover:bg-gray-300">${totalPages}</button>`;
                }
            }

            // Populates a select/dropdown element with options
            function populateSelect(selectId, items, displayKey, valueKey, defaultText = 'Select') {
                const selectElement = document.getElementById(selectId);
                selectElement.innerHTML = `<option value="">${defaultText}</option>`;
                items.forEach(item => {
                    const option = document.createElement('option');
                    option.value = item[valueKey];
                    option.textContent = item[displayKey];
                    selectElement.appendChild(option);
                });
            }
            
            // --- Form Handlers and Event Listeners ---
            // Sets up event handlers for adding, editing, and submitting forms
            function setupForm(model) {
                const singularModel = singularModels[model];

                formButtons[model].addEventListener('click', () => {
                    forms[model].classList.toggle('hidden');
                    document.getElementById(`${singularModel}-form`).reset();
                    document.getElementById(`${singularModel}-id`).value = '';
                    
                    if (model === 'transactions') {
                        const today = new Date().toISOString().split('T')[0];
                        document.getElementById('transaction-transaction-date').value = today;
                        document.getElementById('transaction-tenant').value = '';
                        document.getElementById('transaction-property').value = '';
                        document.getElementById('hidden-transaction-property-id').value = '';
                        document.getElementById('transaction-comments').value = '';
                    }
                    if (model === 'tenants') {
                         document.getElementById('tenant-property').value = '';
                    }
                });

                document.getElementById(`cancel-${singularModel}-btn`).addEventListener('click', () => {
                    forms[model].classList.add('hidden');
                });
                
                document.getElementById(`${singularModel}-form`).addEventListener('submit', async (e) => {
                    e.preventDefault();
                    const form = e.target;
                    const data = Object.fromEntries(new FormData(form).entries());
                    const id = document.getElementById(`${singularModel}-id`).value;
                    
                    for (const key of ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date', 'transaction_date']) {
                        if (data[key]) {
                            data[key] = data[key];
                        }
                    }
                    
                    if (model === 'tenants') {
                        data['property_id'] = parseInt(data['property_id'], 10);
                    }
                    if (model === 'transactions') {
                        data['property_id'] = data['property_id'] ? parseInt(data['property_id'], 10) : null;
                        data['tenant_id'] = data['tenant_id'] ? parseInt(data['tenant_id'], 10) : null;
                    }

                    if (id) {
                        await putData(model, id, data);
                    } else {
                        await postData(model, data);
                    }
                });
            }

            // --- Edit & Delete Handlers ---
            // Sets up event listeners for edit and delete buttons in the tables
            function setupTableHandlers(model) {
                const singularModel = singularModels[model];

                tables[model].addEventListener('click', async (e) => {
                    const row = e.target.closest('tr');
                    if (!row) return;

                    const id = row.dataset.id;
                    
                    if (e.target.classList.contains('delete-btn')) {
                        await deleteData(model, id);
                    } else if (e.target.classList.contains('edit-btn')) {
                        const response = await fetch(`/api/${model}/${id}`);
                        const item = await response.json();
                        
                        forms[model].classList.remove('hidden');
                        document.getElementById(`${singularModel}-id`).value = item.id;
                        
                        if (model === 'tenants') {
                            document.getElementById('tenant-name').value = item.name || '';
                            document.getElementById('tenant-property').value = item.property_id || '';
                            document.getElementById('tenant-passport').value = item.passport || '';
                            document.getElementById('tenant-passport-validity').value = item.passport_validity || '';
                            document.getElementById('tenant-aadhar').value = item.aadhar_no || '';
                            document.getElementById('tenant-employment').value = item.employment_details || '';
                            document.getElementById('tenant-address').value = item.permanent_address || '';
                            document.getElementById('tenant-contact').value = item.contact_no || '';
                            document.getElementById('tenant-emergency-contact').value = item.emergency_contact_no || '';
                            document.getElementById('tenant-rent').value = item.rent || 0;
                            document.getElementById('tenant-security').value = item.security || 0;
                            document.getElementById('tenant-move-in-date').value = item.move_in_date || '';
                            document.getElementById('tenant-contract-start-date').value = item.contract_start_date || '';
                            document.getElementById('tenant-contract-expiry-date').value = item.contract_expiry_date || '';
                        } else if (model === 'properties') {
                            document.getElementById('property-address').value = item.address || '';
                            document.getElementById('property-rent').value = item.rent || 0;
                            document.getElementById('property-maintenance').value = item.maintenance || 0;
                        } else if (model === 'transactions') {
                            document.getElementById('transaction-tenant').value = item.tenant_id || '';
                            document.getElementById('transaction-property').value = item.property_id || '';
                            document.getElementById('hidden-transaction-property-id').value = item.property_id || '';
                            document.getElementById('transaction-type').value = item.type || '';
                            document.getElementById('transaction-for-month').value = item.for_month || '';
                            document.getElementById('transaction-amount').value = item.amount || 0;
                            document.getElementById('transaction-transaction-date').value = item.transaction_date || '';
                            document.getElementById('transaction-comments').value = item.comments || '';
                        }
                    } else if (e.target.classList.contains('tenant-id-link')) {
                        showTenantDetails(e.target.dataset.id);
                    } else if (e.target.classList.contains('property-id-link')) {
                        showPropertyTransactions(e.target.dataset.id);
                    }
                });
            }
            
            // --- CSV Export Handlers ---
            document.getElementById('export-tenants-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/tenants_csv';
            });
            document.getElementById('export-properties-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/properties_csv';
            });
            document.getElementById('export-transactions-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/transactions_csv';
            });

            // --- Transaction Page Specific Handlers ---
            // Handles sorting for the transactions table
            document.querySelectorAll('.sortable-header').forEach(header => {
                header.addEventListener('click', () => {
                    const newSortKey = header.dataset.sortBy;
                    if (sortKey === newSortKey) {
                        sortDirection = sortDirection === 'asc' ? 'desc' : 'asc';
                    } else {
                        sortKey = newSortKey;
                        sortDirection = 'desc';
                    }
                    currentPage = 1;
                    fetchTransactions();
                });
            });
            
            // Handles filtering for the transactions table
            document.getElementById('transaction-type-filter').addEventListener('change', (e) => {
                filterType = e.target.value;
                currentPage = 1;
                fetchTransactions();
            });
            
            document.getElementById('transaction-property-filter').addEventListener('change', (e) => {
                filterPropertyId = e.target.value;
                currentPage = 1;
                fetchTransactions();
            });

            // Handles pagination button clicks
            document.getElementById('prev-page').addEventListener('click', () => {
                if (currentPage > 1) {
                    currentPage--;
                    fetchTransactions();
                }
            });

            document.getElementById('next-page').addEventListener('click', () => {
                if (currentPage < totalPages) {
                    currentPage++;
                    fetchTransactions();
                }
            });
            
            // New logic to auto-fill property on tenant selection
            document.getElementById('transaction-tenant').addEventListener('change', (e) => {
                const selectedTenantId = e.target.value;
                const propertySelect = document.getElementById('transaction-property');
                const hiddenPropertyInput = document.getElementById('hidden-transaction-property-id');
                
                if (selectedTenantId) {
                    const selectedTenant = allTenants.find(tenant => tenant.id == selectedTenantId);
                    if (selectedTenant) {
                        propertySelect.value = selectedTenant.property_id;
                        hiddenPropertyInput.value = selectedTenant.property_id;
                    } else {
                        propertySelect.value = '';
                        hiddenPropertyInput.value = '';
                    }
                } else {
                    propertySelect.value = '';
                    hiddenPropertyInput.value = '';
                }
            });


            // --- Modal Handlers ---
            // Event listeners to close the modals
            document.getElementById('close-tenant-modal-btn').addEventListener('click', () => {
                document.getElementById('tenant-details-modal').classList.remove('visible');
            });
            document.getElementById('close-tenant-modal-footer-btn').addEventListener('click', () => {
                document.getElementById('tenant-details-modal').classList.remove('visible');
            });
            document.getElementById('close-prop-modal-btn').addEventListener('click', () => {
                document.getElementById('property-transactions-modal').classList.remove('visible');
            });
            document.getElementById('close-prop-modal-footer-btn').addEventListener('click', () => {
                document.getElementById('property-transactions-modal').classList.remove('visible');
            });


            // --- Setup all sections ---
            // Initialize the event handlers for all forms and tables
            setupForm('tenants');
            setupForm('properties');
            setupForm('transactions');

            setupTableHandlers('tenants');
            setupTableHandlers('properties');
            setupTableHandlers('transactions');
            
            // Initial data fetch for all sections
            fetchData('tenants');
            fetchData('properties');
        });
    </script>
</body>
</html>
"""

# --- App Configuration ---
app = Flask(__name__)
# The database URI. 'sqlite:///app.db' creates a file named app.db in the app's root directory.
DATABASE_URI = os.getenv('DATABASE_URI', 'sqlite:///app.db')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- Database Models ---
# The schema for our application's data.

class Base(db.Model):
    __abstract__ = True
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(50), default="system")
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_updated_by = db.Column(db.String(50), default="system")

class Tenant(Base):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    property_id = db.Column(db.Integer, db.ForeignKey('property.id'), nullable=True)
    passport = db.Column(db.String(100))
    passport_validity = db.Column(db.Date)
    aadhar_no = db.Column(db.String(100))
    employment_details = db.Column(db.String(255))
    permanent_address = db.Column(db.String(255))
    contact_no = db.Column(db.String(20))
    emergency_contact_no = db.Column(db.String(20))
    rent = db.Column(db.Float, default=0.0)
    security = db.Column(db.Float, default=0.0)
    move_in_date = db.Column(db.Date)
    contract_start_date = db.Column(db.Date)
    contract_expiry_date = db.Column(db.Date)

    # Relationship to Property model
    property = db.relationship('Property', backref='tenants')

    # Converts the model instance to a dictionary for JSON serialization
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'property_id': self.property_id,
            'property_address': self.property.address if self.property else 'N/A',
            'passport': self.passport,
            'passport_validity': self.passport_validity.isoformat() if self.passport_validity else None,
            'aadhar_no': self.aadhar_no,
            'employment_details': self.employment_details,
            'permanent_address': self.permanent_address,
            'contact_no': self.contact_no,
            'emergency_contact_no': self.emergency_contact_no,
            'rent': self.rent,
            'security': self.security,
            'move_in_date': self.move_in_date.isoformat() if self.move_in_date else None,
            'contract_start_date': self.contract_start_date.isoformat() if self.contract_start_date else None,
            'contract_expiry_date': self.contract_expiry_date.isoformat() if self.contract_expiry_date else None,
            'created_date': self.created_date.isoformat() if self.created_date else None,
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat() if self.last_updated else None,
            'last_updated_by': self.last_updated_by
        }

class Property(Base):
    id = db.Column(db.Integer, primary_key=True)
    address = db.Column(db.String(255), nullable=False)
    rent = db.Column(db.Float, default=0.0)
    maintenance = db.Column(db.Float, default=0.0)

    # Converts the model instance to a dictionary for JSON serialization
    def to_dict(self):
        return {
            'id': self.id,
            'address': self.address,
            'rent': self.rent,
            'maintenance': self.maintenance,
            'created_date': self.created_date.isoformat() if self.created_date else None,
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat() if self.last_updated else None,
            'last_updated_by': self.last_updated_by
        }

class Transaction(Base):
    id = db.Column(db.Integer, primary_key=True)
    property_id = db.Column(db.Integer, db.ForeignKey('property.id'), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    type = db.Column(db.String(50), nullable=False)
    for_month = db.Column(db.String(20))
    amount = db.Column(db.Float, nullable=False)
    transaction_date = db.Column(db.Date, default=date.today, nullable=False)
    comments = db.Column(db.String(255))

    # Relationships to Property and Tenant models
    property = db.relationship('Property', backref='transactions')
    tenant = db.relationship('Tenant', backref='transactions')

    # Converts the model instance to a dictionary for JSON serialization
    def to_dict(self):
        return {
            'id': self.id,
            'property_id': self.property_id,
            'tenant_id': self.tenant_id,
            'property_address': self.property.address if self.property else 'N/A',
            'tenant_name': self.tenant.name if self.tenant else 'N/A',
            'type': self.type,
            'for_month': self.for_month,
            'amount': self.amount,
            'created_date': self.created_date.isoformat() if self.created_date else None,
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat() if self.last_updated else None,
            'last_updated_by': self.last_updated_by,
            'transaction_date': self.transaction_date.isoformat() if self.transaction_date else None,
            'comments': self.comments
        }

# --- API Endpoints ---
# These endpoints handle the business logic and data interaction.

@app.route('/api/<string:model>', methods=['GET', 'POST'])
def api_list(model):
    """Handles GET (list all) and POST (create new) requests for all models."""
    if model == 'tenants':
        Model = Tenant
    elif model == 'properties':
        Model = Property
    elif model == 'transactions':
        Model = Transaction
    else:
        return jsonify({'error': 'Invalid model'}), 400

    if request.method == 'POST':
        try:
            data = request.json
            # Type casting for incoming JSON data
            if model == 'tenants':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['security'] = float(data.get('security', 0) or 0)
                if data.get('property_id'):
                    data['property_id'] = int(data['property_id'])
                else:
                    data['property_id'] = None
                for date_field in ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date']:
                    if data.get(date_field):
                        data[date_field] = datetime.strptime(data[date_field], '%Y-%m-%d').date()
                    else:
                        data[date_field] = None
            elif model == 'properties':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['maintenance'] = float(data.get('maintenance', 0) or 0)
            elif model == 'transactions':
                data['property_id'] = int(data.get('property_id'))
                data['tenant_id'] = int(data.get('tenant_id')) if data.get('tenant_id') else None
                data['amount'] = float(data.get('amount', 0) or 0)
                if data.get('transaction_date'):
                    data['transaction_date'] = datetime.strptime(data['transaction_date'], '%Y-%m-%d').date()
                else:
                    data['transaction_date'] = date.today()
                data['comments'] = data.get('comments', None)

            instance = Model(**data)
            db.session.add(instance)
            db.session.commit()
            return jsonify(instance.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Invalid data or required field missing: {str(e)}'}), 400
    else: # GET
        if model == 'transactions':
            # Pagination, filtering, and sorting for transactions
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 50, type=int)
            filter_type = request.args.get('type', 'all', type=str)
            filter_property_id = request.args.get('property_id', None, type=str)
            sort_by = request.args.get('sort_by', 'transaction_date', type=str)
            sort_direction = request.args.get('sort_direction', 'desc', type=str)

            query = db.session.query(Transaction)

            if filter_type != 'all':
                query = query.filter(Transaction.type == filter_type)
            if filter_property_id:
                query = query.filter(Transaction.property_id == filter_property_id)

            if sort_direction == 'asc':
                query = query.order_by(getattr(Transaction, sort_by).asc())
            else:
                query = query.order_by(getattr(Transaction, sort_by).desc())
            
            pagination = db.paginate(query, page=page, per_page=per_page, error_out=False)
            
            return jsonify({
                'items': [item.to_dict() for item in pagination.items],
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total_pages': pagination.pages,
                'total_items': pagination.total
            })
        else:
            instances = Model.query.all()
            return jsonify([instance.to_dict() for instance in instances])

@app.route('/api/<string:model>/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def api_detail(model, id):
    """Handles GET (read), PUT (update), and DELETE (delete) requests for a single record."""
    if model == 'tenants':
        Model = Tenant
    elif model == 'properties':
        Model = Property
    elif model == 'transactions':
        Model = Transaction
    else:
        return jsonify({'error': 'Invalid model'}), 400
    
    instance = Model.query.get_or_404(id)

    if request.method == 'GET':
        return jsonify(instance.to_dict())
    
    elif request.method == 'PUT':
        try:
            data = request.json
            # Type casting for incoming JSON data
            if model == 'tenants':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['security'] = float(data.get('security', 0) or 0)
                if data.get('property_id'):
                    data['property_id'] = int(data['property_id'])
                else:
                    data['property_id'] = None
                for date_field in ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date']:
                    if data.get(date_field):
                        data[date_field] = datetime.strptime(data[date_field], '%Y-%m-%d').date()
                    else:
                        data[date_field] = None
            elif model == 'properties':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['maintenance'] = float(data.get('maintenance', 0) or 0)
            elif model == 'transactions':
                data['property_id'] = int(data.get('property_id'))
                data['tenant_id'] = int(data.get('tenant_id')) if data.get('tenant_id') else None
                data['amount'] = float(data.get('amount', 0) or 0)
                if data.get('transaction_date'):
                    data['transaction_date'] = datetime.strptime(data['transaction_date'], '%Y-%m-%d').date()
                else:
                    data.pop('transaction_date', None)
                data['comments'] = data.get('comments', None)

            for key, value in data.items():
                setattr(instance, key, value)
            db.session.commit()
            return jsonify(instance.to_dict()), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Invalid data or required field missing: {str(e)}'}), 400
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(instance)
            db.session.commit()
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 400

@app.route('/api/properties/<int:id>/transactions', methods=['GET'])
def get_property_transactions(id):
    """Fetches all transactions for a specific property and calculates the total balance."""
    property_instance = Property.query.get_or_404(id)
    
    transactions = Transaction.query.filter_by(property_id=id).order_by(Transaction.transaction_date.desc()).all()
    
    transactions_list = [tx.to_dict() for tx in transactions]
    
    # Calculate the total balance for the property
    total_balance = 0.0
    for tx in transactions:
        if tx.type == 'payment_received':
            total_balance += tx.amount
        else:
            total_balance -= tx.amount
    
    return jsonify({
        'transactions': transactions_list,
        'total': total_balance
    })


# --- Report Generation Endpoints (Excel) ---

def generate_excel_report(data, headers, title):
    """Helper function to create an Excel file from data."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True)
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_data)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/api/reports/tenants')
def report_tenants_xlsx():
    """Generates and downloads an Excel report of all tenants."""
    tenants = Tenant.query.all()
    headers = [
        'ID', 'Name', 'Property Address', 'Passport', 'Passport Validity', 'Aadhar No', 'Employment Details',
        'Permanent Address', 'Contact No', 'Emergency Contact No', 'Rent', 'Security',
        'Move In Date', 'Contract Start Date', 'Contract Expiry Date', 'Created Date'
    ]
    data = [
        [
            t.id, t.name, t.property.address if t.property else 'N/A', t.passport, t.passport_validity, t.aadhar_no, t.employment_details,
            t.permanent_address, t.contact_no, t.emergency_contact_no, t.rent, t.security,
            t.move_in_date, t.contract_start_date, t.contract_expiry_date, t.created_date
        ] for t in tenants
    ]
    report_file = generate_excel_report(data, headers, "Tenants Report")
    return send_file(
        report_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='tenants_report.xlsx'
    )

@app.route('/api/reports/transactions')
def report_transactions_xlsx():
    """Generates and downloads an Excel report of all transactions."""
    transactions = Transaction.query.all()
    headers = [
        'ID', 'Property Address', 'Tenant Name', 'Type', 'For Month', 'Amount', 'Transaction Date', 'Comments'
    ]
    data = [
        [
            t.id, t.property.address if t.property else 'N/A',
            t.tenant.name if t.tenant else 'N/A',
            t.type, t.for_month, t.amount, t.transaction_date, t.comments
        ] for t in transactions
    ]
    report_file = generate_excel_report(data, headers, "Transactions Report")
    return send_file(
        report_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='transactions_report.xlsx'
    )
    
# --- CSV Export Endpoints ---

def generate_csv_report(data, headers):
    """Helper function to create a CSV file from data."""
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(headers)
    cw.writerows(data)
    output = BytesIO(si.getvalue().encode('utf-8'))
    return output

@app.route('/api/reports/tenants_csv')
def report_tenants_csv():
    """Generates and downloads a CSV report of all tenants."""
    tenants = Tenant.query.all()
    headers = [
        'ID', 'Name', 'Property Address', 'Passport', 'Passport Validity', 'Aadhar No', 'Employment Details',
        'Permanent Address', 'Contact No', 'Emergency Contact No', 'Rent', 'Security',
        'Move In Date', 'Contract Start Date', 'Contract Expiry Date', 'Created Date'
    ]
    data = [
        [
            t.id, t.name, t.property.address if t.property else 'N/A', t.passport, t.passport_validity, t.aadhar_no, t.employment_details,
            t.permanent_address, t.contact_no, t.emergency_contact_no, t.rent, t.security,
            t.move_in_date, t.contract_start_date, t.contract_expiry_date, t.created_date
        ] for t in tenants
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='tenants_report.csv'
    )

@app.route('/api/reports/properties_csv')
def report_properties_csv():
    """Generates and downloads a CSV report of all properties."""
    properties = Property.query.all()
    headers = [
        'ID', 'Address', 'Rent', 'Maintenance', 'Created Date'
    ]
    data = [
        [
            p.id, p.address, p.rent, p.maintenance, p.created_date
        ] for p in properties
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='properties_report.csv'
    )

@app.route('/api/reports/transactions_csv')
def report_transactions_csv():
    """Generates and downloads a CSV report of all transactions."""
    transactions = Transaction.query.all()
    headers = [
        'ID', 'Property Address', 'Tenant Name', 'Type', 'For Month', 'Amount', 'Transaction Date', 'Comments'
    ]
    data = [
        [
            t.id, t.property.address if t.property else 'N/A',
            t.tenant.name if t.tenant else 'N/A',
            t.type, t.for_month, t.amount, t.transaction_date, t.comments
        ] for t in transactions
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='transactions_report.csv'
    )

@app.route('/api/backup')
def backup_database():
    """
    Creates a timestamped backup of the database file on the server
    and sends it to the user for download.
    """
    try:
        # Get the database path from the configured DATABASE_URI
        database_uri = app.config['SQLALCHEMY_DATABASE_URI']
        
        # Handle SQLite database paths
        if database_uri.startswith('sqlite:///'):
            # Remove 'sqlite:///' prefix
            db_filename = database_uri.replace('sqlite:///', '')
            
            # Flask creates an 'instance' directory for the database
            # Construct the full path: app_root/instance/db_filename
            db_path = os.path.join(app.root_path, 'instance', db_filename)
        else:
            return jsonify({'error': 'Backup is only supported for SQLite databases'}), 400
        
        # Check if the database file exists before attempting to back it up
        if not os.path.exists(db_path):
            return jsonify({'error': f'Database file not found at: {db_path}. Please create some records first to generate the database.'}), 404
        
        # Get backup storage path from environment variable, default to current directory
        backup_storage_path = os.getenv('BACKUP_STORAGE_PATH', '.')
        
        # Ensure the backup directory exists
        os.makedirs(backup_storage_path, exist_ok=True)
        
        # Create a dynamic filename for the backup
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"app_backup_{timestamp}.db"
        
        # Create the full backup file path
        backup_file_path = os.path.join(backup_storage_path, backup_filename)
        
        # Save a copy on the server's file system
        shutil.copy2(db_path, backup_file_path)
        
        # Send the newly created backup file to the user
        return send_file(
            backup_file_path,
            as_attachment=True,
            download_name=backup_filename,
            mimetype='application/octet-stream'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/')
def index():
    """Renders the main HTML page for the application."""
    return render_template_string(HTML_TEMPLATE)

if __name__ == '__main__':
    with app.app_context():
        # This will create the database tables if they don't already exist
        db.create_all()
    app.run(debug=True)
