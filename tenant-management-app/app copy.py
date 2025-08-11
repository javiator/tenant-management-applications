# app.py - A single-file, full-stack property management application.

import os
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO, StringIO
import csv

# --- Embedded Frontend HTML/CSS/JS ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Property Management</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;700&display=swap');
        body {
            font-family: 'Inter', sans-serif;
            background-color: #f3f4f6;
        }
        .tab-button {
            transition: all 0.2s;
            font-weight: 500;
        }
        .tab-button.active {
            border-bottom-width: 2px;
            border-color: #3b82f6;
            color: #3b82f6;
        }
        .main-container {
            max-width: 1200px;
        }
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.75);
            display: flex;
            justify-content: center;
            align-items: center;
            z-index: 1000;
            transition: opacity 0.3s ease-in-out;
            opacity: 0;
            pointer-events: none;
        }
        .modal-overlay.visible {
            opacity: 1;
            pointer-events: auto;
        }
        .modal-content {
            background-color: #fff;
            padding: 2rem;
            border-radius: 1rem;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.2);
            max-width: 90%;
            max-height: 90vh;
            overflow-y: auto;
            transform: scale(0.9);
            transition: transform 0.3s ease-in-out;
        }
        .modal-overlay.visible .modal-content {
            transform: scale(1);
        }
        .id-link {
            cursor: pointer;
            color: #3b82f6;
            text-decoration: underline;
        }
        .sortable-header {
            cursor: pointer;
            user-select: none;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .sortable-header:hover {
            color: #1e40af;
        }
        .sort-icon {
            width: 1rem;
            height: 1rem;
            margin-left: 0.25rem;
        }
        .pagination-button.active {
            background-color: #3b82f6;
            color: #fff;
        }
        .pagination-button:disabled {
            cursor: not-allowed;
            opacity: 0.5;
        }
    </style>
</head>
<body class="bg-gray-100 flex items-center justify-center min-h-screen p-4">
    <div class="bg-white shadow-xl rounded-2xl w-full main-container">
        <!-- Header -->
        <header class="p-6 border-b border-gray-200">
            <h1 class="text-3xl font-bold text-gray-800">Property Management</h1>
            <p class="text-gray-500 mt-1">Manage tenants, properties, and transactions with ease.</p>
        </header>

        <!-- Tabs Navigation -->
        <nav class="flex border-b border-gray-200 text-sm md:text-base px-6">
            <button id="tab-tenants" class="tab-button py-4 px-4 text-gray-600 active focus:outline-none">Tenants</button>
            <button id="tab-properties" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Properties</button>
            <button id="tab-transactions" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Transactions</button>
            <button id="tab-reports" class="tab-button py-4 px-4 text-gray-600 focus:outline-none">Reports</button>
        </nav>

        <main class="p-6">
            <!-- Tenants Section -->
            <section id="section-tenants" class="tab-content">
                <div class="flex justify-between items-center mb-6">
                    <h2 class="text-xl font-bold text-gray-800">Tenants</h2>
                    <div class="flex space-x-2">
                        <button id="export-tenants-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-tenant-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Tenant</button>
                    </div>
                </div>
                <div id="tenant-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="tenant-form" class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4">
                        <!-- All tenant input fields here -->
                        <input type="hidden" id="tenant-id">
                        <div>
                            <label for="tenant-name" class="block text-sm font-medium text-gray-700">Name</label>
                            <input type="text" id="tenant-name" name="name" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-passport" class="block text-sm font-medium text-gray-700">Passport</label>
                            <input type="text" id="tenant-passport" name="passport" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-passport-validity" class="block text-sm font-medium text-gray-700">Passport Validity</label>
                            <input type="date" id="tenant-passport-validity" name="passport_validity" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-aadhar" class="block text-sm font-medium text-gray-700">Aadhar No</label>
                            <input type="text" id="tenant-aadhar" name="aadhar_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-employment" class="block text-sm font-medium text-gray-700">Employment Details</label>
                            <input type="text" id="tenant-employment" name="employment_details" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-address" class="block text-sm font-medium text-gray-700">Permanent Address</label>
                            <input type="text" id="tenant-address" name="permanent_address" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contact" class="block text-sm font-medium text-gray-700">Contact No</label>
                            <input type="text" id="tenant-contact" name="contact_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-emergency-contact" class="block text-sm font-medium text-gray-700">Emergency Contact No</label>
                            <input type="text" id="tenant-emergency-contact" name="emergency_contact_no" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-rent" class="block text-sm font-medium text-gray-700">Rent</label>
                            <input type="number" id="tenant-rent" name="rent" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-security" class="block text-sm font-medium text-gray-700">Security Deposit</label>
                            <input type="number" id="tenant-security" name="security" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-move-in-date" class="block text-sm font-medium text-gray-700">Move In Date</label>
                            <input type="date" id="tenant-move-in-date" name="move_in_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contract-start-date" class="block text-sm font-medium text-gray-700">Contract Start Date</label>
                            <input type="date" id="tenant-contract-start-date" name="contract_start_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="tenant-contract-expiry-date" class="block text-sm font-medium text-gray-700">Contract Expiry Date</label>
                            <input type="date" id="tenant-contract-expiry-date" name="contract_expiry_date" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2 lg:col-span-3 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-tenant-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Tenant</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">ID</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Name</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Contact No</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Rent</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Move In Date</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="tenant-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Tenant data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
            </section>

            <!-- Properties Section -->
            <section id="section-properties" class="tab-content hidden">
                <div class="flex justify-between items-center mb-6">
                    <h2 class="text-xl font-bold text-gray-800">Properties</h2>
                    <div class="flex space-x-2">
                        <button id="export-properties-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-property-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Property</button>
                    </div>
                </div>
                <div id="property-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="property-form" class="grid grid-cols-1 md:grid-cols-2 gap-4">
                        <input type="hidden" id="property-id">
                        <div>
                            <label for="property-address" class="block text-sm font-medium text-gray-700">Address</label>
                            <input type="text" id="property-address" name="address" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="property-rent" class="block text-sm font-medium text-gray-700">Rent</label>
                            <input type="number" id="property-rent" name="rent" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="property-maintenance" class="block text-sm font-medium text-gray-700">Maintenance</label>
                            <input type="number" id="property-maintenance" name="maintenance" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-property-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Property</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">ID</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Address</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Rent</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Maintenance</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="property-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Property data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
            </section>

            <!-- Transactions Section -->
            <section id="section-transactions" class="tab-content hidden">
                <div class="flex flex-col md:flex-row justify-between items-start md:items-center mb-6 space-y-4 md:space-y-0">
                    <h2 class="text-xl font-bold text-gray-800">Transactions</h2>
                    <div class="flex flex-col sm:flex-row space-y-4 sm:space-y-0 sm:space-x-4">
                        <div>
                            <label for="transaction-property-filter" class="text-sm font-medium text-gray-700 mr-2">Filter by Property:</label>
                            <select id="transaction-property-filter" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></select>
                        </div>
                        <div>
                            <label for="transaction-type-filter" class="text-sm font-medium text-gray-700 mr-2">Filter by Type:</label>
                            <select id="transaction-type-filter" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="all">All</option>
                                <option value="rent">Rent</option>
                                <option value="security">Security</option>
                                <option value="payment_received">Payment Received</option>
                                <option value="gas">Gas</option>
                                <option value="electricity">Electricity</option>
                                <option value="water">Water</option>
                                <option value="maintenance">Maintenance</option>
                                <option value="misc">Miscellaneous</option>
                            </select>
                        </div>
                        <button id="export-transactions-csv" class="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Export CSV</button>
                        <button id="add-transaction-btn" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Add Transaction</button>
                    </div>
                </div>
                <div id="transaction-form-container" class="bg-gray-50 p-4 rounded-xl mb-6 hidden">
                    <form id="transaction-form" class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4">
                        <input type="hidden" id="transaction-id">
                        <div>
                            <label for="transaction-property" class="block text-sm font-medium text-gray-700">Property</label>
                            <select id="transaction-property" name="property_id" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></select>
                        </div>
                        <div>
                            <label for="transaction-tenant" class="block text-sm font-medium text-gray-700">Tenant</label>
                            <select id="transaction-tenant" name="tenant_id" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500"></select>
                        </div>
                        <div>
                            <label for="transaction-type" class="block text-sm font-medium text-gray-700">Type</label>
                            <select id="transaction-type" name="type" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                                <option value="rent">Rent</option>
                                <option value="security">Security</option>
                                <option value="payment_received">Payment Received</option>
                                <option value="gas">Gas</option>
                                <option value="electricity">Electricity</option>
                                <option value="water">Water</option>
                                <option value="maintenance">Maintenance</option>
                                <option value="misc">Miscellaneous</option>
                            </select>
                        </div>
                        <div>
                            <label for="transaction-for-month" class="block text-sm font-medium text-gray-700">For Month</label>
                            <input type="text" id="transaction-for-month" name="for_month" class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div>
                            <label for="transaction-amount" class="block text-sm font-medium text-gray-700">Amount</label>
                            <input type="number" id="transaction-amount" name="amount" required class="mt-1 block w-full rounded-md border-gray-300 shadow-sm focus:border-blue-500 focus:ring-blue-500">
                        </div>
                        <div class="col-span-1 md:col-span-2 lg:col-span-3 flex justify-end space-x-2 mt-4">
                            <button type="button" id="cancel-transaction-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300 transition-colors duration-200">Cancel</button>
                            <button type="submit" class="bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 shadow-md">Save Transaction</button>
                        </div>
                    </form>
                </div>
                <div class="overflow-x-auto bg-white rounded-lg shadow-sm">
                    <table class="min-w-full divide-y divide-gray-200">
                        <thead class="bg-gray-50">
                            <tr>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="id">
                                        <span>ID</span>
                                        <span class="sort-icon" id="sort-icon-id"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Property</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Tenant</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="type">
                                        <span>Type</span>
                                        <span class="sort-icon" id="sort-icon-type"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">For Month</th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="amount">
                                        <span>Amount</span>
                                        <span class="sort-icon" id="sort-icon-amount"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">
                                    <div class="sortable-header" data-sort-by="created_date">
                                        <span>Date</span>
                                        <span class="sort-icon" id="sort-icon-created_date"></span>
                                    </div>
                                </th>
                                <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Actions</th>
                            </tr>
                        </thead>
                        <tbody id="transaction-table-body" class="bg-white divide-y divide-gray-200">
                            <!-- Transaction data will be populated here by JavaScript -->
                        </tbody>
                    </table>
                </div>
                <div id="pagination-controls" class="flex justify-center items-center space-x-2 mt-6">
                    <button id="prev-page" class="pagination-button bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Previous</button>
                    <div id="page-numbers" class="flex space-x-1"></div>
                    <button id="next-page" class="pagination-button bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Next</button>
                </div>
            </section>

            <!-- Reports Section -->
            <section id="section-reports" class="tab-content hidden">
                <h2 class="text-xl font-bold text-gray-800 mb-4">Reports</h2>
                <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
                    <div class="bg-gray-50 p-6 rounded-xl shadow-sm flex flex-col items-start">
                        <h3 class="text-lg font-semibold text-gray-800">All Tenants Report (xlsx)</h3>
                        <p class="text-gray-600 mt-2">Generate a detailed report of all tenant information in an Excel file.</p>
                        <a href="/api/reports/tenants" class="mt-4 bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 self-end">Download .xlsx</a>
                    </div>
                    <div class="bg-gray-50 p-6 rounded-xl shadow-sm flex flex-col items-start">
                        <h3 class="text-lg font-semibold text-gray-800">All Transactions Report (xlsx)</h3>
                        <p class="text-gray-600 mt-2">Generate a detailed report of all transactions in an Excel file.</p>
                        <a href="/api/reports/transactions" class="mt-4 bg-blue-600 hover:bg-blue-700 text-white font-bold py-2 px-4 rounded-full transition-colors duration-200 self-end">Download .xlsx</a>
                    </div>
                </div>
            </section>
        </main>
    </div>

    <!-- Tenant Details Modal -->
    <div id="tenant-details-modal" class="modal-overlay">
        <div class="modal-content">
            <div class="flex justify-between items-center pb-4 border-b mb-4">
                <h3 class="text-2xl font-bold text-gray-800">Tenant Details</h3>
                <button id="close-tenant-modal-btn" class="text-gray-400 hover:text-gray-700 text-3xl">&times;</button>
            </div>
            <div id="tenant-details-content" class="grid grid-cols-1 md:grid-cols-2 gap-4 text-gray-700">
                <!-- Details will be populated here by JavaScript -->
            </div>
            <div class="flex justify-end mt-6">
                <button id="close-tenant-modal-footer-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Close</button>
            </div>
        </div>
    </div>
    
    <!-- Property Transactions Modal -->
    <div id="property-transactions-modal" class="modal-overlay">
        <div class="modal-content w-full max-w-4xl">
            <div class="flex justify-between items-center pb-4 border-b mb-4">
                <h3 class="text-2xl font-bold text-gray-800">Transactions for Property #<span id="prop-id-in-modal"></span></h3>
                <button id="close-prop-modal-btn" class="text-gray-400 hover:text-gray-700 text-3xl">&times;</button>
            </div>
            <div class="overflow-x-auto">
                <table class="min-w-full divide-y divide-gray-200">
                    <thead class="bg-gray-50">
                        <tr>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Tenant</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Type</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">For Month</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Amount</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Date</th>
                        </tr>
                    </thead>
                    <tbody id="transactions-modal-table-body" class="bg-white divide-y divide-gray-200">
                        <!-- Transaction data will be populated here by JavaScript -->
                    </tbody>
                </table>
            </div>
            <div id="transactions-modal-footer" class="mt-6 flex justify-between items-center">
                <div class="text-lg font-bold text-gray-800">
                    Total (last 15 transactions): <span id="transactions-modal-total" class="text-blue-600"></span>
                </div>
                <button id="close-prop-modal-footer-btn" class="bg-gray-200 text-gray-700 font-bold py-2 px-4 rounded-full hover:bg-gray-300">Close</button>
            </div>
        </div>
    </div>

    <script>
        document.addEventListener('DOMContentLoaded', () => {
            const tabs = document.querySelectorAll('.tab-button');
            const sections = document.querySelectorAll('.tab-content');
            
            // Map plural model names to singular for consistent ID access
            const singularModels = {
                'tenants': 'tenant',
                'properties': 'property',
                'transactions': 'transaction'
            };

            const forms = {
                'tenants': document.getElementById('tenant-form-container'),
                'properties': document.getElementById('property-form-container'),
                'transactions': document.getElementById('transaction-form-container'),
            };
            const formButtons = {
                'tenants': document.getElementById('add-tenant-btn'),
                'properties': document.getElementById('add-property-btn'),
                'transactions': document.getElementById('add-transaction-btn'),
            };
            const tables = {
                'tenants': document.getElementById('tenant-table-body'),
                'properties': document.getElementById('property-table-body'),
                'transactions': document.getElementById('transaction-table-body'),
            };
            
            // Global data caches and state variables
            let allProperties = [];
            let allTenants = [];
            let allTransactions = []; // This will now only store the currently viewed page
            
            // State for the transactions page
            let currentPage = 1;
            const itemsPerPage = 50; // Updated to 50 items per page
            let totalPages = 1;
            let filterType = 'all';
            let filterPropertyId = 'all';
            let sortKey = 'created_date';
            let sortDirection = 'desc';

            let currentActiveTab = 'tenants';

            function showSection(sectionId) {
                sections.forEach(section => {
                    section.classList.add('hidden');
                });
                document.getElementById(`section-${sectionId}`).classList.remove('hidden');

                tabs.forEach(tab => {
                    tab.classList.remove('active', 'text-blue-600', 'border-blue-600');
                    tab.classList.add('text-gray-600');
                });
                document.getElementById(`tab-${sectionId}`).classList.add('active', 'text-blue-600', 'border-blue-600');
                currentActiveTab = sectionId;
            }

            tabs.forEach(tab => {
                tab.addEventListener('click', () => {
                    const sectionId = tab.id.split('-')[1];
                    showSection(sectionId);
                    if (sectionId !== 'reports') {
                        // For transactions, we need to handle pagination and filtering
                        if (sectionId === 'transactions') {
                            currentPage = 1; // Reset to page 1 on tab switch
                            fetchTransactions();
                        } else {
                            fetchData(sectionId);
                        }
                    }
                });
            });
            
            // Initial load
            showSection('tenants');
            fetchData('tenants');
            fetchData('properties');

            // --- API Functions ---
            async function fetchData(model) {
                const response = await fetch(`/api/${model}`);
                const data = await response.json();
                
                if (model === 'tenants') {
                    allTenants = data;
                    renderTable(model, data);
                } else if (model === 'properties') {
                    allProperties = data;
                    renderTable(model, data);
                    // This is the fix for the filter dropdown: ensure properties are fetched before populating the select
                    populateSelect('transaction-property-filter', allProperties, 'address', 'id', 'All');
                    populateSelect('transaction-property', allProperties, 'address', 'id', '-- Select Property --');
                } else {
                    renderTable(model, data);
                }
            }
            
            async function fetchTransactions() {
                // Ensure properties and tenants are fetched if not already in cache
                if (allProperties.length === 0) {
                    await fetchData('properties');
                }
                if (allTenants.length === 0) {
                    await fetchData('tenants');
                }
                
                const url = `/api/transactions?page=${currentPage}&per_page=${itemsPerPage}&type=${filterType}&property_id=${filterPropertyId}&sort_by=${sortKey}&sort_direction=${sortDirection}`;
                const response = await fetch(url);
                const data = await response.json();
                
                allTransactions = data.items;
                totalPages = data.total_pages;
                
                renderTransactionsTable();
                renderPaginationControls();
            }
            
            async function postData(model, data) {
                const response = await fetch(`/api/${model}`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                if (response.ok) {
                    forms[model].classList.add('hidden');
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error creating record: ${error.error}`);
                }
            }

            async function putData(model, id, data) {
                const response = await fetch(`/api/${model}/${id}`, {
                    method: 'PUT',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                if (response.ok) {
                    forms[model].classList.add('hidden');
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error updating record: ${error.error}`);
                }
            }

            async function deleteData(model, id) {
                if (!confirm(`Are you sure you want to delete this ${singularModels[model]}?`)) return;
                
                const response = await fetch(`/api/${model}/${id}`, {
                    method: 'DELETE'
                });
                if (response.ok) {
                    if (model === 'transactions') {
                        fetchTransactions();
                    } else {
                        fetchData(model);
                    }
                } else {
                    const error = await response.json();
                    alert(`Error deleting record: ${error.error}`);
                }
            }

            async function showTenantDetails(id) {
                try {
                    const response = await fetch(`/api/tenants/${id}`);
                    if (!response.ok) throw new Error('Tenant not found.');
                    const tenant = await response.json();
                    
                    const detailsContent = document.getElementById('tenant-details-content');
                    detailsContent.innerHTML = `
                        <div class="col-span-1 md:col-span-2">
                            <p class="font-semibold">Name:</p>
                            <p>${tenant.name}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Passport:</p>
                            <p>${tenant.passport || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Passport Validity:</p>
                            <p>${tenant.passport_validity || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Aadhar No:</p>
                            <p>${tenant.aadhar_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Employment Details:</p>
                            <p>${tenant.employment_details || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Permanent Address:</p>
                            <p>${tenant.permanent_address || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contact No:</p>
                            <p>${tenant.contact_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Emergency Contact No:</p>
                            <p>${tenant.emergency_contact_no || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Rent:</p>
                            <p>$${(tenant.rent || 0).toFixed(2)}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Security Deposit:</p>
                            <p>$${(tenant.security || 0).toFixed(2)}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Move In Date:</p>
                            <p>${tenant.move_in_date || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contract Start Date:</p>
                            <p>${tenant.contract_start_date || 'N/A'}</p>
                        </div>
                        <div>
                            <p class="font-semibold">Contract Expiry Date:</p>
                            <p>${tenant.contract_expiry_date || 'N/A'}</p>
                        </div>
                    `;
                    document.getElementById('tenant-details-modal').classList.add('visible');
                } catch (error) {
                    alert('Error fetching tenant details.');
                }
            }

            async function showPropertyTransactions(id) {
                try {
                    const response = await fetch(`/api/properties/${id}/transactions`);
                    if (!response.ok) throw new Error('Could not fetch transactions.');
                    const { transactions, total } = await response.json();
                    
                    document.getElementById('prop-id-in-modal').textContent = id;
                    const transactionsTableBody = document.getElementById('transactions-modal-table-body');
                    transactionsTableBody.innerHTML = '';
                    
                    if (transactions.length === 0) {
                        transactionsTableBody.innerHTML = '<tr><td colspan="5" class="text-center py-4 text-gray-500">No recent transactions found.</td></tr>';
                    } else {
                        transactions.forEach(tx => {
                            const date = new Date(tx.created_date).toLocaleDateString();
                            transactionsTableBody.innerHTML += `
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.tenant_name || 'N/A'}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.type}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${tx.for_month || 'N/A'}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${tx.amount.toFixed(2)}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${date}</td>
                                </tr>
                            `;
                        });
                    }
                    
                    document.getElementById('transactions-modal-total').textContent = `$${total.toFixed(2)}`;
                    document.getElementById('property-transactions-modal').classList.add('visible');
                } catch (error) {
                    alert('Error fetching property transactions.');
                }
            }
            
            // --- Rendering Functions ---
            function renderTable(model, data) {
                tables[model].innerHTML = '';
                let rowsHtml = '';
                if (model === 'tenants') {
                    data.forEach(item => {
                        rowsHtml += `
                            <tr data-id="${item.id}">
                                <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">
                                    <button class="id-link tenant-id-link" data-id="${item.id}">${item.id}</button>
                                </td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.name}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.contact_no || 'N/A'}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.rent.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.move_in_date || 'N/A'}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                    <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                    <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                                </td>
                            </tr>
                        `;
                    });
                } else if (model === 'properties') {
                    data.forEach(item => {
                        rowsHtml += `
                            <tr data-id="${item.id}">
                                <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">
                                    <button class="id-link property-id-link" data-id="${item.id}">${item.id}</button>
                                </td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.address}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.rent.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.maintenance.toFixed(2)}</td>
                                <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                    <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                    <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                                </td>
                            </tr>
                        `;
                    });
                }
                tables[model].innerHTML = rowsHtml;
            }

            function renderTransactionsTable() {
                tables['transactions'].innerHTML = '';
                let rowsHtml = '';
                allTransactions.forEach(item => {
                    const date = new Date(item.created_date).toLocaleDateString();
                    rowsHtml += `
                        <tr data-id="${item.id}">
                            <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">${item.id}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.property_address || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.tenant_name || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.type}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${item.for_month || 'N/A'}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">$${item.amount.toFixed(2)}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-500">${date}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                <button class="edit-btn text-indigo-600 hover:text-indigo-900 mr-2">Edit</button>
                                <button class="delete-btn text-red-600 hover:text-red-900">Delete</button>
                            </td>
                        </tr>
                    `;
                });
                tables['transactions'].innerHTML = rowsHtml;
                updateSortIcons();
            }

            function updateSortIcons() {
                document.querySelectorAll('.sort-icon').forEach(icon => {
                    icon.innerHTML = '';
                });
                if (sortKey) {
                    const iconElement = document.getElementById(`sort-icon-${sortKey}`);
                    if (iconElement) {
                        iconElement.innerHTML = sortDirection === 'asc' ? '&#9650;' : '&#9660;'; // Up/down arrow
                    }
                }
            }
            
            function renderPaginationControls() {
                const paginationDiv = document.getElementById('pagination-controls');
                const pageNumbersDiv = document.getElementById('page-numbers');
                const prevButton = document.getElementById('prev-page');
                const nextButton = document.getElementById('next-page');
                
                pageNumbersDiv.innerHTML = '';
                
                prevButton.disabled = currentPage === 1;
                nextButton.disabled = currentPage === totalPages || totalPages === 0;

                const startPage = Math.max(1, currentPage - 2);
                const endPage = Math.min(totalPages, currentPage + 2);
                
                if (startPage > 1) {
                    pageNumbersDiv.innerHTML += `<button class="pagination-button py-2 px-3 rounded-full hover:bg-gray-300">1</button><span>...</span>`;
                }

                for (let i = startPage; i <= endPage; i++) {
                    const button = document.createElement('button');
                    button.textContent = i;
                    button.classList.add('pagination-button', 'py-2', 'px-3', 'rounded-full', 'hover:bg-gray-300');
                    if (i === currentPage) {
                        button.classList.add('active');
                    }
                    button.addEventListener('click', () => {
                        currentPage = i;
                        fetchTransactions();
                    });
                    pageNumbersDiv.appendChild(button);
                }
                
                if (endPage < totalPages) {
                    pageNumbersDiv.innerHTML += `<span>...</span><button class="pagination-button py-2 px-3 rounded-full hover:bg-gray-300">${totalPages}</button>`;
                }
            }


            function populateSelect(selectId, items, displayKey, valueKey, defaultText = 'Select') {
                const selectElement = document.getElementById(selectId);
                selectElement.innerHTML = `<option value="all">${defaultText}</option>`;
                items.forEach(item => {
                    const option = document.createElement('option');
                    option.value = item[valueKey];
                    option.textContent = item[displayKey];
                    selectElement.appendChild(option);
                });
            }
            
            // --- Form Handlers and Event Listeners ---
            function setupForm(model) {
                const singularModel = singularModels[model];

                formButtons[model].addEventListener('click', () => {
                    forms[model].classList.toggle('hidden');
                    document.getElementById(`${singularModel}-form`).reset();
                    document.getElementById(`${singularModel}-id`).value = '';
                });

                document.getElementById(`cancel-${singularModel}-btn`).addEventListener('click', () => {
                    forms[model].classList.add('hidden');
                });
                
                document.getElementById(`${singularModel}-form`).addEventListener('submit', async (e) => {
                    e.preventDefault();
                    const form = e.target;
                    const data = Object.fromEntries(new FormData(form).entries());
                    const id = document.getElementById(`${singularModel}-id`).value;
                    
                    // Convert date fields to ISO format for the backend
                    for (const key of ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date']) {
                        if (data[key]) {
                            // Backend expects 'YYYY-MM-DD'
                            data[key] = data[key];
                        }
                    }

                    if (id) {
                        await putData(model, id, data);
                    } else {
                        await postData(model, data);
                    }
                });
            }

            // --- Edit & Delete Handlers ---
            function setupTableHandlers(model) {
                const singularModel = singularModels[model];

                tables[model].addEventListener('click', async (e) => {
                    const row = e.target.closest('tr');
                    if (!row) return;

                    const id = row.dataset.id;
                    
                    if (e.target.classList.contains('delete-btn')) {
                        await deleteData(model, id);
                    } else if (e.target.classList.contains('edit-btn')) {
                        const response = await fetch(`/api/${model}/${id}`);
                        const item = await response.json();
                        
                        forms[model].classList.remove('hidden');
                        document.getElementById(`${singularModel}-id`).value = item.id;
                        
                        // Populate form fields
                        for (const key in item) {
                            const input = document.getElementById(`${singularModel}-${key.replace(/_/g, '-')}`);
                            if (input) {
                                input.value = item[key] || '';
                            }
                        }
                    } else if (e.target.classList.contains('tenant-id-link')) {
                        showTenantDetails(e.target.dataset.id);
                    } else if (e.target.classList.contains('property-id-link')) {
                        showPropertyTransactions(e.target.dataset.id);
                    }
                });
            }
            
            // --- CSV Export Handlers ---
            document.getElementById('export-tenants-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/tenants_csv';
            });
            document.getElementById('export-properties-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/properties_csv';
            });
            document.getElementById('export-transactions-csv').addEventListener('click', () => {
                window.location.href = '/api/reports/transactions_csv';
            });

            // --- Transaction Page Specific Handlers ---
            document.querySelectorAll('.sortable-header').forEach(header => {
                header.addEventListener('click', () => {
                    const newSortKey = header.dataset.sortBy;
                    if (sortKey === newSortKey) {
                        sortDirection = sortDirection === 'asc' ? 'desc' : 'asc';
                    } else {
                        sortKey = newSortKey;
                        sortDirection = 'desc'; // Default to descending for date
                    }
                    currentPage = 1;
                    fetchTransactions();
                });
            });
            
            document.getElementById('transaction-type-filter').addEventListener('change', (e) => {
                filterType = e.target.value;
                currentPage = 1;
                fetchTransactions();
            });
            
            document.getElementById('transaction-property-filter').addEventListener('change', (e) => {
                filterPropertyId = e.target.value;
                currentPage = 1;
                fetchTransactions();
            });

            document.getElementById('prev-page').addEventListener('click', () => {
                if (currentPage > 1) {
                    currentPage--;
                    fetchTransactions();
                }
            });

            document.getElementById('next-page').addEventListener('click', () => {
                if (currentPage < totalPages) {
                    currentPage++;
                    fetchTransactions();
                }
            });
            
            document.getElementById('transaction-property').addEventListener('change', async (e) => {
                const propertyId = e.target.value;
                const tenantSelect = document.getElementById('transaction-tenant');
                
                // Reset tenant dropdown
                tenantSelect.innerHTML = '<option value="">-- Select Tenant --</option>';
                
                if (propertyId && propertyId !== 'all') {
                    // Fetch all transactions for this specific property
                    const response = await fetch(`/api/properties/${propertyId}/transactions`);
                    const data = await response.json();
                    
                    const tenantIdsForProperty = new Set(data.transactions.map(tx => tx.tenant_id));
                    
                    // Filter the global list of tenants
                    const filteredTenants = allTenants.filter(tenant => tenantIdsForProperty.has(tenant.id));
                    
                    // Populate the dropdown with filtered tenants
                    populateSelect('transaction-tenant', filteredTenants, 'name', 'id', '-- Select Tenant --');
                } else {
                    // If no property is selected, show all tenants
                    populateSelect('transaction-tenant', allTenants, 'name', 'id', '-- Select Tenant --');
                }
            });

            // --- Modal Handlers ---
            document.getElementById('close-tenant-modal-btn').addEventListener('click', () => {
                document.getElementById('tenant-details-modal').classList.remove('visible');
            });
            document.getElementById('close-tenant-modal-footer-btn').addEventListener('click', () => {
                document.getElementById('tenant-details-modal').classList.remove('visible');
            });
            document.getElementById('close-prop-modal-btn').addEventListener('click', () => {
                document.getElementById('property-transactions-modal').classList.remove('visible');
            });
            document.getElementById('close-prop-modal-footer-btn').addEventListener('click', () => {
                document.getElementById('property-transactions-modal').classList.remove('visible');
            });


            // --- Setup all sections ---
            setupForm('tenants');
            setupForm('properties');
            setupForm('transactions');

            setupTableHandlers('tenants');
            setupTableHandlers('properties');
            setupTableHandlers('transactions');
            
            // Initial data fetches for transactions and properties
            fetchData('tenants');
            fetchData('properties');
        });
    </script>
</body>
</html>
"""

# --- App Configuration ---
app = Flask(__name__)
# Configure a SQLite database, which is stored in a file named 'app.db'.
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///app.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- Database Models ---

# Base model with common fields for tracking creation and updates.
class Base(db.Model):
    __abstract__ = True
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(50), default="system")
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_updated_by = db.Column(db.String(50), default="system")

# Tenant Model
class Tenant(Base):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    passport = db.Column(db.String(100))
    passport_validity = db.Column(db.Date)
    aadhar_no = db.Column(db.String(100))
    employment_details = db.Column(db.String(255))
    permanent_address = db.Column(db.String(255))
    contact_no = db.Column(db.String(20))
    emergency_contact_no = db.Column(db.String(20))
    rent = db.Column(db.Float, default=0.0)
    security = db.Column(db.Float, default=0.0)
    move_in_date = db.Column(db.Date)
    contract_start_date = db.Column(db.Date)
    contract_expiry_date = db.Column(db.Date)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'passport': self.passport,
            'passport_validity': self.passport_validity.isoformat() if self.passport_validity else None,
            'aadhar_no': self.aadhar_no,
            'employment_details': self.employment_details,
            'permanent_address': self.permanent_address,
            'contact_no': self.contact_no,
            'emergency_contact_no': self.emergency_contact_no,
            'rent': self.rent,
            'security': self.security,
            'move_in_date': self.move_in_date.isoformat() if self.move_in_date else None,
            'contract_start_date': self.contract_start_date.isoformat() if self.contract_start_date else None,
            'contract_expiry_date': self.contract_expiry_date.isoformat() if self.contract_expiry_date else None,
            'created_date': self.created_date.isoformat(),
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat(),
            'last_updated_by': self.last_updated_by
        }

# Property Model
class Property(Base):
    id = db.Column(db.Integer, primary_key=True)
    address = db.Column(db.String(255), nullable=False)
    rent = db.Column(db.Float, default=0.0)
    maintenance = db.Column(db.Float, default=0.0)

    def to_dict(self):
        return {
            'id': self.id,
            'address': self.address,
            'rent': self.rent,
            'maintenance': self.maintenance,
            'created_date': self.created_date.isoformat(),
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat(),
            'last_updated_by': self.last_updated_by
        }

# Transaction Model
class Transaction(Base):
    id = db.Column(db.Integer, primary_key=True)
    property_id = db.Column(db.Integer, db.ForeignKey('property.id'), nullable=False)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'))
    type = db.Column(db.String(50), nullable=False)  # e.g., 'rent', 'security', 'gas', 'electricity'
    for_month = db.Column(db.String(20))
    amount = db.Column(db.Float, nullable=False)

    property = db.relationship('Property', backref='transactions')
    tenant = db.relationship('Tenant', backref='transactions')

    def to_dict(self):
        return {
            'id': self.id,
            'property_id': self.property_id,
            'tenant_id': self.tenant_id,
            'property_address': self.property.address if self.property else 'N/A',
            'tenant_name': self.tenant.name if self.tenant else 'N/A',
            'type': self.type,
            'for_month': self.for_month,
            'amount': self.amount,
            'created_date': self.created_date.isoformat(),
            'created_by': self.created_by,
            'last_updated': self.last_updated.isoformat(),
            'last_updated_by': self.last_updated_by
        }

# --- API Endpoints ---

@app.route('/')
def index():
    # The HTML, CSS, and JS for the frontend are embedded as a single string.
    # This makes the app a single, self-contained file.
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/<string:model>', methods=['GET', 'POST'])
def api_list(model):
    if model == 'tenants':
        Model = Tenant
    elif model == 'properties':
        Model = Property
    elif model == 'transactions':
        Model = Transaction
    else:
        return jsonify({'error': 'Invalid model'}), 400

    if request.method == 'POST':
        try:
            data = request.json
            if model == 'tenants':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['security'] = float(data.get('security', 0) or 0)
                for date_field in ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date']:
                    if data.get(date_field):
                        data[date_field] = datetime.strptime(data[date_field], '%Y-%m-%d').date()
                    else:
                        data[date_field] = None
            elif model == 'properties':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['maintenance'] = float(data.get('maintenance', 0) or 0)
            elif model == 'transactions':
                data['property_id'] = int(data.get('property_id'))
                data['tenant_id'] = int(data.get('tenant_id')) if data.get('tenant_id') else None
                data['amount'] = float(data.get('amount', 0) or 0)

            instance = Model(**data)
            db.session.add(instance)
            db.session.commit()
            return jsonify(instance.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Invalid data or required field missing: {str(e)}'}), 400
    else: # GET
        if model == 'transactions':
            # Handle pagination and filtering for transactions
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 50, type=int)
            filter_type = request.args.get('type', 'all', type=str)
            filter_property_id = request.args.get('property_id', 'all', type=str)
            sort_by = request.args.get('sort_by', 'created_date', type=str)
            sort_direction = request.args.get('sort_direction', 'desc', type=str)

            query = db.session.query(Transaction)

            if filter_type != 'all':
                query = query.filter(Transaction.type == filter_type)
            if filter_property_id != 'all':
                query = query.filter(Transaction.property_id == filter_property_id)

            if sort_direction == 'asc':
                query = query.order_by(getattr(Transaction, sort_by).asc())
            else:
                query = query.order_by(getattr(Transaction, sort_by).desc())
            
            # Use `db.paginate` for modern Flask-SQLAlchemy pagination
            pagination = db.paginate(query, page=page, per_page=per_page, error_out=False)
            
            return jsonify({
                'items': [item.to_dict() for item in pagination.items],
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total_pages': pagination.pages,
                'total_items': pagination.total
            })
        else:
            # Standard GET for other models
            instances = Model.query.all()
            return jsonify([instance.to_dict() for instance in instances])

@app.route('/api/<string:model>/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def api_detail(model, id):
    if model == 'tenants':
        Model = Tenant
    elif model == 'properties':
        Model = Property
    elif model == 'transactions':
        Model = Transaction
    else:
        return jsonify({'error': 'Invalid model'}), 400
    
    instance = Model.query.get_or_404(id)

    if request.method == 'GET':
        return jsonify(instance.to_dict())
    
    elif request.method == 'PUT':
        try:
            data = request.json
            if model == 'tenants':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['security'] = float(data.get('security', 0) or 0)
                for date_field in ['passport_validity', 'move_in_date', 'contract_start_date', 'contract_expiry_date']:
                    if data.get(date_field):
                        data[date_field] = datetime.strptime(data[date_field], '%Y-%m-%d').date()
                    else:
                        data[date_field] = None
            elif model == 'properties':
                data['rent'] = float(data.get('rent', 0) or 0)
                data['maintenance'] = float(data.get('maintenance', 0) or 0)
            elif model == 'transactions':
                data['property_id'] = int(data.get('property_id'))
                data['tenant_id'] = int(data.get('tenant_id')) if data.get('tenant_id') else None
                data['amount'] = float(data.get('amount', 0) or 0)

            for key, value in data.items():
                setattr(instance, key, value)
            db.session.commit()
            return jsonify(instance.to_dict()), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Invalid data or required field missing: {str(e)}'}), 400
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(instance)
            db.session.commit()
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 400

# New API endpoint to get the last 15 transactions for a property.
@app.route('/api/properties/<int:id>/transactions', methods=['GET'])
def get_property_transactions(id):
    property_instance = Property.query.get_or_404(id)
    
    # Get the last 15 transactions for this property, ordered by creation date
    transactions = Transaction.query.filter_by(property_id=id).order_by(Transaction.created_date.desc()).limit(15).all()
    
    transactions_list = [tx.to_dict() for tx in transactions]
    total_amount = sum(tx.amount for tx in transactions)
    
    return jsonify({
        'transactions': transactions_list,
        'total': total_amount
    })


# --- Report Generation Endpoints (Excel) ---

def generate_excel_report(data, headers, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Header styling
    header_font = Font(bold=True)
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_data)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/api/reports/tenants')
def report_tenants_xlsx():
    tenants = Tenant.query.all()
    headers = [
        'ID', 'Name', 'Passport', 'Passport Validity', 'Aadhar No', 'Employment Details',
        'Permanent Address', 'Contact No', 'Emergency Contact No', 'Rent', 'Security',
        'Move In Date', 'Contract Start Date', 'Contract Expiry Date', 'Created Date'
    ]
    data = [
        [
            t.id, t.name, t.passport, t.passport_validity, t.aadhar_no, t.employment_details,
            t.permanent_address, t.contact_no, t.emergency_contact_no, t.rent, t.security,
            t.move_in_date, t.contract_start_date, t.contract_expiry_date, t.created_date
        ] for t in tenants
    ]
    report_file = generate_excel_report(data, headers, "Tenants Report")
    return send_file(
        report_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='tenants_report.xlsx'
    )

@app.route('/api/reports/transactions')
def report_transactions_xlsx():
    transactions = Transaction.query.all()
    headers = [
        'ID', 'Property Address', 'Tenant Name', 'Type', 'For Month', 'Amount', 'Created Date'
    ]
    data = [
        [
            t.id, t.property.address if t.property else 'N/A',
            t.tenant.name if t.tenant else 'N/A',
            t.type, t.for_month, t.amount, t.created_date
        ] for t in transactions
    ]
    report_file = generate_excel_report(data, headers, "Transactions Report")
    return send_file(
        report_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='transactions_report.xlsx'
    )
    
# --- New CSV Export Endpoints ---

def generate_csv_report(data, headers):
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(headers)
    cw.writerows(data)
    output = BytesIO(si.getvalue().encode('utf-8'))
    return output

@app.route('/api/reports/tenants_csv')
def report_tenants_csv():
    tenants = Tenant.query.all()
    headers = [
        'ID', 'Name', 'Passport', 'Passport Validity', 'Aadhar No', 'Employment Details',
        'Permanent Address', 'Contact No', 'Emergency Contact No', 'Rent', 'Security',
        'Move In Date', 'Contract Start Date', 'Contract Expiry Date', 'Created Date'
    ]
    data = [
        [
            t.id, t.name, t.passport, t.passport_validity, t.aadhar_no, t.employment_details,
            t.permanent_address, t.contact_no, t.emergency_contact_no, t.rent, t.security,
            t.move_in_date, t.contract_start_date, t.contract_expiry_date, t.created_date
        ] for t in tenants
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='tenants_report.csv'
    )

@app.route('/api/reports/properties_csv')
def report_properties_csv():
    properties = Property.query.all()
    headers = [
        'ID', 'Address', 'Rent', 'Maintenance', 'Created Date'
    ]
    data = [
        [
            p.id, p.address, p.rent, p.maintenance, p.created_date
        ] for p in properties
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='properties_report.csv'
    )

@app.route('/api/reports/transactions_csv')
def report_transactions_csv():
    transactions = Transaction.query.all()
    headers = [
        'ID', 'Property Address', 'Tenant Name', 'Type', 'For Month', 'Amount', 'Created Date'
    ]
    data = [
        [
            t.id, t.property.address if t.property else 'N/A',
            t.tenant.name if t.tenant else 'N/A',
            t.type, t.for_month, t.amount, t.created_date
        ] for t in transactions
    ]
    report_file = generate_csv_report(data, headers)
    return send_file(
        report_file,
        mimetype='text/csv',
        as_attachment=True,
        download_name='transactions_report.csv'
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
